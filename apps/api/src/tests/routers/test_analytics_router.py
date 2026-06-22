"""Router tests for src/routers/analytics.py."""

from contextlib import contextmanager
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, Mock, patch

import httpx
import pytest
from fastapi import FastAPI, HTTPException
from httpx import ASGITransport, AsyncClient

import src.routers.analytics as analytics_router_module
from src.core.events.database import get_db_session
from src.db.users import AnonymousUser
from src.routers.analytics import (
    _enrich_with_metadata,
    _execute_tinybird_query,
    _get_read_client,
    _parse_safe_params,
    _validate_course_uuid,
    _verify_org_admin,
    _verify_org_membership,
    router as analytics_router,
)
from src.security.auth import get_current_user


@pytest.fixture
def db_session():
    session = AsyncMock()
    result = MagicMock()
    result.all.return_value = []
    result.first.return_value = None
    scalars = MagicMock()
    scalars.all.return_value = []
    scalars.first.return_value = None
    result.scalars.return_value = scalars
    session.execute.return_value = result
    return session


@pytest.fixture
def current_user():
    return SimpleNamespace(id=1)


@pytest.fixture
def app(db_session, current_user):
    app = FastAPI()
    app.include_router(analytics_router, prefix="/api/v1/analytics")
    app.dependency_overrides[get_db_session] = lambda: db_session
    app.dependency_overrides[get_current_user] = lambda: current_user
    yield app
    app.dependency_overrides.clear()


@pytest.fixture
async def client(app):
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as c:
        yield c


@contextmanager
def _analytics_guard_patches():
    with (
        patch("src.routers.analytics._verify_org_membership", new_callable=AsyncMock),
        patch("src.routers.analytics._verify_org_admin", new_callable=AsyncMock),
    ):
        yield


def _result(all_result=None, first_result=None):
    result = Mock()
    result.all.return_value = [] if all_result is None else all_result
    result.first.return_value = first_result
    return result


class TestAnalyticsHelpers:
    def test_validate_course_uuid_and_sql_helpers(self):
        request = SimpleNamespace(query_params={"days": "14"})
        assert _parse_safe_params(3, request, 30) == (3, 14)
        assert _validate_course_uuid("course_1") == "course_1"
        assert _validate_course_uuid("c" * 100) == "c" * 100

        with pytest.raises(HTTPException):
            _parse_safe_params("bad", request, 30)

        with pytest.raises(HTTPException):
            _validate_course_uuid("bad course")

        with pytest.raises(HTTPException):
            _validate_course_uuid("c" * 101)

    def test_get_read_client_handles_missing_and_configured_tinybird(self, monkeypatch):
        monkeypatch.setattr(analytics_router_module, "_read_client", None)
        monkeypatch.setattr(
            "src.routers.analytics.get_learnhouse_config",
            lambda: SimpleNamespace(tinybird_config=None),
        )
        assert _get_read_client() is None

        fake_client = SimpleNamespace()
        monkeypatch.setattr(analytics_router_module, "_read_client", None)
        monkeypatch.setattr(
            "src.routers.analytics.get_learnhouse_config",
            lambda: SimpleNamespace(
                tinybird_config=SimpleNamespace(
                    api_url="https://tinybird.test",
                    read_token="token",
                )
            ),
        )
        with patch("src.routers.analytics.httpx.AsyncClient", return_value=fake_client):
            assert _get_read_client() is fake_client

        monkeypatch.setattr(analytics_router_module, "_read_client", fake_client)
        assert _get_read_client() is fake_client

    async def test_verify_org_membership_and_admin(self):
        db = AsyncMock()

        def _make_execute_result(first_result=None, all_result=None):
            scalars = MagicMock()
            scalars.first.return_value = first_result
            scalars.all.return_value = all_result if all_result is not None else []
            result = MagicMock()
            result.scalars.return_value = scalars
            return result

        with patch("src.routers.analytics.is_user_superadmin", new_callable=AsyncMock, return_value=True):
            await _verify_org_membership(1, 10, db)
            await _verify_org_admin(1, 10, db)

        db.execute.return_value = _make_execute_result(first_result=None)
        with patch("src.routers.analytics.is_user_superadmin", new_callable=AsyncMock, return_value=False):
            with pytest.raises(HTTPException, match="Not a member"):
                await _verify_org_membership(1, 10, db)

        membership = SimpleNamespace(role_id=7)
        db.execute.side_effect = [
            _make_execute_result(first_result=membership),
            _make_execute_result(first_result=SimpleNamespace(rights={"organizations": {"action_update": True}})),
        ]
        with patch("src.routers.analytics.is_user_superadmin", new_callable=AsyncMock, return_value=False):
            await _verify_org_admin(1, 10, db)

        db.execute.side_effect = [
            _make_execute_result(first_result=membership),
            _make_execute_result(first_result=SimpleNamespace(rights={"organizations": {"action_update": False}})),
        ]
        with patch("src.routers.analytics.is_user_superadmin", new_callable=AsyncMock, return_value=False):
            with pytest.raises(HTTPException, match="Admin access required"):
                await _verify_org_admin(1, 10, db)

        db.execute.side_effect = [_make_execute_result(first_result=None)]
        with patch("src.routers.analytics.is_user_superadmin", new_callable=AsyncMock, return_value=False):
            with pytest.raises(HTTPException, match="Admin access required"):
                await _verify_org_admin(1, 10, db)

        db.execute.side_effect = [
            _make_execute_result(first_result=membership),
            _make_execute_result(first_result=None),
        ]
        with patch("src.routers.analytics.is_user_superadmin", new_callable=AsyncMock, return_value=False):
            with pytest.raises(HTTPException, match="Admin access required"):
                await _verify_org_admin(1, 10, db)

    async def test_enrich_with_metadata(self):
        db = AsyncMock()

        def _make_execute_result(all_result=None):
            scalars = MagicMock()
            scalars.all.return_value = all_result if all_result is not None else []
            result = MagicMock()
            result.scalars.return_value = scalars
            return result

        rows = [
            {"user_id": 9, "course_uuid": "course_1", "activity_uuid": "activity_1", "last_activity_uuid": "activity_2"},
            {"user_id": 10, "activity_uuid": "activity_2"},
        ]
        courses = [
            SimpleNamespace(id=1, course_uuid="course_1", name="Course 1", thumbnail_image=None),
            SimpleNamespace(id=2, course_uuid="course_2", name="Course 2", thumbnail_image="thumb"),
        ]
        activities = [
            SimpleNamespace(activity_uuid="activity_1", name="Activity 1", course_id=1),
            SimpleNamespace(activity_uuid="activity_2", name="Activity 2", course_id=2),
        ]
        db.execute.side_effect = [
            _make_execute_result(all_result=courses),
            _make_execute_result(all_result=activities),
            _make_execute_result(all_result=[courses[1]]),
        ]

        enriched = await _enrich_with_metadata(rows, db)

        assert enriched[0]["course_name"] == "Course 1"
        assert enriched[0]["thumbnail_image"] == ""
        assert enriched[0]["activity_name"] == "Activity 1"
        assert enriched[0]["last_activity_name"] == "Activity 2"
        assert enriched[1]["course_uuid"] == "course_2"
        assert enriched[1]["course_name"] == "Course 2"

    async def test_enrich_with_metadata_empty_and_missing_course_resolution(self):
        db = AsyncMock()
        assert await _enrich_with_metadata([], db) == []

        def _make_execute_result(all_result=None):
            scalars = MagicMock()
            scalars.all.return_value = all_result if all_result is not None else []
            result = MagicMock()
            result.scalars.return_value = scalars
            return result

        rows = [{"activity_uuid": "activity_1"}]
        db.execute.side_effect = [
            _make_execute_result(all_result=[SimpleNamespace(activity_uuid="activity_1", name="Activity 1", course_id=2)]),
            _make_execute_result(all_result=[SimpleNamespace(id=2, course_uuid="course_2", name="Course 2", thumbnail_image="thumb")]),
        ]

        enriched = await _enrich_with_metadata(rows, db)

        assert enriched[0]["course_uuid"] == "course_2"
        assert enriched[0]["course_name"] == "Course 2"
        assert enriched[0]["activity_name"] == "Activity 1"

    def test_build_sql_validates_parameter_types_and_course_uuid(self):
        template = "SELECT {org_id} AS org_id, {days} AS days"

        assert analytics_router_module._build_sql(template, 1, 7) == "SELECT 1 AS org_id, 7 AS days"

        with pytest.raises(HTTPException, match="Invalid parameter types"):
            analytics_router_module._build_sql(template, "bad", 7)

        with pytest.raises(HTTPException, match="Invalid course_uuid"):
            analytics_router_module._build_sql(template, 1, 7, "bad course")

    @pytest.mark.asyncio
    async def test_execute_tinybird_query_cache_and_error_paths(self):
        cached = {"data": [{"ok": True}], "rows": 1, "meta": []}
        with patch("src.routers.analytics.get_cached_result", return_value=cached):
            assert await _execute_tinybird_query("daily_active_users", "sql", 1, 30) == cached

        with patch("src.routers.analytics.get_cached_result", return_value=None), patch(
            "src.routers.analytics._get_read_client",
            return_value=None,
        ):
            with pytest.raises(HTTPException, match="Analytics not configured"):
                await _execute_tinybird_query("daily_active_users", "sql", 1, 30)

        request = httpx.Request("POST", "https://tinybird.test/v0/sql")
        not_found_response = httpx.Response(404, request=request, content=b"table not found")
        not_found_exc = httpx.HTTPStatusError("not found", request=request, response=not_found_response)
        fake_client = SimpleNamespace(post=AsyncMock(side_effect=not_found_exc))
        with patch("src.routers.analytics.get_cached_result", return_value=None), patch(
            "src.routers.analytics._get_read_client",
            return_value=fake_client,
        ), patch("src.routers.analytics.set_cached_result") as set_cached:
            result = await _execute_tinybird_query("daily_active_users", "sql", 1, 30)
        assert result == {"data": [], "rows": 0, "meta": []}

        bad_response = httpx.Response(500, request=request, content=b"boom")
        bad_exc = httpx.HTTPStatusError("boom", request=request, response=bad_response)
        fake_client = SimpleNamespace(post=AsyncMock(side_effect=bad_exc))
        with patch("src.routers.analytics.get_cached_result", return_value=None), patch(
            "src.routers.analytics._get_read_client",
            return_value=fake_client,
        ):
            with pytest.raises(HTTPException, match="Analytics query failed"):
                await _execute_tinybird_query("daily_active_users", "sql", 1, 30)

        fake_client = SimpleNamespace(post=AsyncMock(side_effect=RuntimeError("boom")))
        with patch("src.routers.analytics.get_cached_result", return_value=None), patch(
            "src.routers.analytics._get_read_client",
            return_value=fake_client,
        ):
            with pytest.raises(HTTPException, match="Analytics query failed"):
                await _execute_tinybird_query("daily_active_users", "sql", 1, 30)

        payload = {
            "data": [{"value": float("nan"), "other": float("inf"), "negative": float("-inf"), "ok": 1}],
            "rows": 1,
            "meta": [{"field": "value"}],
        }
        fake_response = SimpleNamespace(
            raise_for_status=lambda: None,
            json=lambda: payload,
        )
        fake_client = SimpleNamespace(post=AsyncMock(return_value=fake_response))
        with patch("src.routers.analytics.get_cached_result", return_value=None), patch(
            "src.routers.analytics._get_read_client",
            return_value=fake_client,
        ), patch("src.routers.analytics.set_cached_result") as set_cached:
            result = await _execute_tinybird_query("daily_active_users", "sql", 1, 30)
        assert result["data"][0]["value"] is None
        assert result["data"][0]["other"] is None
        assert result["data"][0]["negative"] is None
        set_cached.assert_called_once()

    def test_write_xlsx_sheet_empty_rows(self):
        from openpyxl import Workbook
        from src.routers.analytics import _write_xlsx_sheet

        wb = Workbook()
        ws = wb.active
        _write_xlsx_sheet(ws, [])

        assert ws.cell(1, 1).value == "No data"
        assert ws.cell(2, 1).value is None


    def test_write_xlsx_sheet_non_bool_and_bool_columns(self):
        from openpyxl import Workbook
        from src.routers.analytics import (
            _write_xlsx_sheet, _BOOL_HIGHLIGHT_COLS, _GRN_FILL, _RED_FILL,
        )

        wb = Workbook()
        ws = wb.active
        bool_col = next(iter(_BOOL_HIGHLIGHT_COLS))
        rows = [
            {"Name": "Alice", bool_col: "Sì"},
            {"Name": "Bob",   bool_col: "No"},
        ]
        _write_xlsx_sheet(ws, rows)

        # colonna non-bool: nessun fill colorato
        assert ws.cell(2, 1).fill != _GRN_FILL
        assert ws.cell(2, 1).fill != _RED_FILL
        # colonna bool: fill verde/rosso
        assert ws.cell(2, 2).fill == _GRN_FILL
        assert ws.cell(3, 2).fill == _RED_FILL


    def test_build_xlsx_course_users_structure(self):
        from src.routers.analytics import _build_xlsx_course_users

        flat = [
            {"Groups": "Alpha",       "Subscribed": "Yes", "Completed course": "Yes", "Name": "Alice"},
            {"Groups": "Alpha, Beta", "Subscribed": "Yes", "Completed course": "No",  "Name": "Bob"},
            {"Groups": "None",        "Subscribed": "No",  "Completed course": "No",  "Name": "Carlo"},
        ]
        wb = _build_xlsx_course_users(flat)

        assert "Summary"   in wb.sheetnames
        assert "All"       in wb.sheetnames
        assert "Alpha"     in wb.sheetnames
        assert "Beta"      in wb.sheetnames
        assert "No groups" in wb.sheetnames


    def test_build_xlsx_course_users_sheet_name_truncated(self):
        from src.routers.analytics import _build_xlsx_course_users

        long_name = "A" * 40
        flat = [{"Groups": long_name, "Subscribed": "No", "Completed course": "No", "Name": "X"}]
        wb = _build_xlsx_course_users(flat)

        assert long_name[:31] in wb.sheetnames


    def test_flat_to_csv_with_data_and_empty(self):
        from src.routers.analytics import _flat_to_csv

        flat = [{"Name": "Alice", "Subscribed": "Yes"}, {"Name": "Bob", "Subscribed": "No"}]
        result = _flat_to_csv(flat)
        assert "Name,Subscribed" in result
        assert "Alice,Yes"       in result
        assert "Bob,No"          in result

        assert _flat_to_csv([]) == ""
        
    def test_build_sql_with_limit_valid_and_invalid(self):
        template = "SELECT {org_id} AS org_id, {days} AS days, {limit} AS limit"

        # limit valido
        result = analytics_router_module._build_sql(template, 1, 7, limit=10)
        assert result == "SELECT 1 AS org_id, 7 AS days, 10 AS limit"

        # limit = 0 → invalid
        with pytest.raises(HTTPException, match="Invalid limit"):
            analytics_router_module._build_sql(template, 1, 7, limit=0)

        # limit negativo → invalid
        with pytest.raises(HTTPException, match="Invalid limit"):
            analytics_router_module._build_sql(template, 1, 7, limit=-1)

    def test_sanitize_function(self):
        from src.routers.analytics import _sanitize

        # list e dict → stringa
        assert _sanitize([1, 2, 3]) == "[1, 2, 3]"
        assert _sanitize({"a": 1}) == "{'a': 1}"

        # float NaN/Inf → None
        assert _sanitize(float("nan")) is None
        assert _sanitize(float("inf")) is None
        assert _sanitize(float("-inf")) is None

        # valori normali → invariati
        assert _sanitize("hello") == "hello"
        assert _sanitize(42) == 42
        assert _sanitize(None) is None

    def test_write_learners_sheet_with_data(self):
        from openpyxl import Workbook
        from src.routers.analytics import _write_learners_sheet

        results = {
            "daily_active_users": {"data": [{"date": "2024-01-01", "dau": 10}]},
            "new_vs_returning":   {"data": [{"date": "2024-01-01", "new_users": 5, "returning_users": 3}]},
            "cohort_retention":   {"data": [{"cohort_week": "2024-W01", "cohort_size": 20, "week_1": 15, "week_2": 10, "week_4": 8, "week_8": 5}]},
            "peak_usage_hours":   {"data": [{"day_of_week": 1, "event_count": 100}, {"day_of_week": 2, "event_count": 80}]},
        }
        wb = Workbook()
        ws = wb.active
        _write_learners_sheet(ws, results)
        # verifica che la sheet abbia contenuto
        assert ws.cell(1, 1).value == "Learners & Growth"

    def test_write_learners_sheet_empty_data(self):
        from openpyxl import Workbook
        from src.routers.analytics import _write_learners_sheet

        # tutti i dataset vuoti → branch "else: No data" per ogni sezione
        results = {
            "daily_active_users": {"data": []},
            "new_vs_returning":   {"data": []},
            "cohort_retention":   {"data": []},
            "peak_usage_hours":   {"data": []},
        }
        wb = Workbook()
        ws = wb.active
        _write_learners_sheet(ws, results)
        assert ws.cell(1, 1).value == "Learners & Growth"

    def test_write_courses_sheet_with_and_without_data(self):
        from openpyxl import Workbook
        from src.routers.analytics import _write_courses_sheet

        # con dati → branch "if rows"
        results_with_data = {
            "top_courses":                 {"data": [{"course_name": "Course A", "views": 10, "enrollments": 5, "completions": 2}]},
            "course_rating_by_completion": {"data": []},
            "time_to_completion":          {"data": []},
            "content_type_effectiveness":  {"data": []},
            "course_dropoff":              {"data": []},
            "completion_velocity":         {"data": []},
        }
        wb = Workbook()
        ws = wb.active
        _write_courses_sheet(ws, results_with_data)
        assert ws.cell(1, 1).value == "Courses & Content"

        # tutti vuoti → branch "else: No data" per ogni sezione
        results_empty = {k: {"data": []} for k in results_with_data}
        wb2 = Workbook()
        ws2 = wb2.active
        _write_courses_sheet(ws2, results_empty)
        assert ws2.cell(1, 1).value == "Courses & Content"

    def test_json_course_export_structure(self):
        from src.routers.analytics import _json_course_export

        flat = [
            {"Subscribed": "Yes", "Completed Course": "Yes", "Name": "Alice"},
            {"Subscribed": "Yes", "Completed Course": "No",  "Name": "Bob"},
            {"Subscribed": "No",  "Completed Course": "No",  "Name": "Carlo"},
        ]
        result = _json_course_export(flat, org_id=1, days=30, course_uuid="course_1")

        assert result["meta"]["org_id"] == 1
        assert result["meta"]["total_users"] == 3
        assert result["meta"]["enrolled"] == 2
        assert result["meta"]["completed"] == 1
        assert result["meta"]["completion_rate"] == 50.0
        assert result["users"] == flat

        # lista vuota → completion_rate None
        empty = _json_course_export([], org_id=1, days=30, course_uuid="course_1")
        assert empty["meta"]["completion_rate"] is None

    def test_results_to_csv_includes_query_name_and_skips_unknown(self):
        from src.routers.analytics import _results_to_csv

        results = {
            "daily_active_users": {"data": [{"date": "2024-01-01", "dau": 5}]},
            "unknown_query":      {"data": [{"foo": "bar"}]},          # non in _CSV_ORG_QUERIES → skippato
            "enrollment_funnel":  {"data": []},                        # in _CSV_ORG_QUERIES ma vuoto
        }
        csv_text = _results_to_csv(results)

        assert "# daily_active_users" in csv_text
        assert "unknown_query" not in csv_text
        assert "# enrollment_funnel" in csv_text
        assert "# No data" in csv_text

    async def test_export_with_course_uuid_json_csv_xlsx(self, client):
        flat_mock = [
            {
                "ID": 1, "Username": "alice", "First Name": "Alice", "Last Name": "Smith",
                "Email": "alice@test.com", "Role": "student", "Signup Method": "email",
                "Verified Mail": "Sì", "Last Login": "", "Signup Date": "",
                "Groups": "None", "Subscribed": "Yes", "Subscription Date": "",
                "Completed Course": "Yes", "Completion Date": "",
                "Completed Activities": 5, "Total Seconds": 3600,
            }
        ]
        with _analytics_guard_patches(), patch(
            "src.routers.analytics._fetch_course_users_flat",
            new_callable=AsyncMock,
            return_value=flat_mock,
        ):
            response = await client.get(
                "/api/v1/analytics/export?org_id=1&format=json&course_uuid=course_1"
            )
        assert response.status_code == 200
        assert response.json()["meta"]["course_uuid"] == "course_1"
        assert response.json()["meta"]["enrolled"] == 1

        with _analytics_guard_patches(), patch(
            "src.routers.analytics._fetch_course_users_flat",
            new_callable=AsyncMock,
            return_value=flat_mock,
        ):
            response = await client.get(
                "/api/v1/analytics/export?org_id=1&format=csv&course_uuid=course_1"
            )
        assert response.status_code == 200
        assert "Alice" in response.text

        with _analytics_guard_patches(), patch(
            "src.routers.analytics._fetch_course_users_flat",
            new_callable=AsyncMock,
            return_value=flat_mock,
        ):
            response = await client.get(
                "/api/v1/analytics/export?org_id=1&format=xlsx&course_uuid=course_1"
            )
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class TestAnalyticsRouter:
    async def test_status_and_plan_info(self, client, app):
        with patch("src.routers.analytics.get_learnhouse_config", return_value=SimpleNamespace(tinybird_config=None)):
            response = await client.get("/api/v1/analytics/status")
        assert response.status_code == 200
        assert response.json()["configured"] is False

        app.dependency_overrides[get_current_user] = lambda: AnonymousUser()
        response = await client.get("/api/v1/analytics/status")
        assert response.status_code == 401
        app.dependency_overrides[get_current_user] = lambda: SimpleNamespace(id=1)

        with patch("src.routers.analytics.get_org_plan", new_callable=AsyncMock, return_value="pro"), patch(
            "src.routers.analytics.plan_meets_requirement",
            return_value=True,
        ), patch("src.routers.analytics._verify_org_membership", new_callable=AsyncMock), patch(
            "src.routers.analytics._verify_org_admin", new_callable=AsyncMock,
        ):
            response = await client.get("/api/v1/analytics/plan-info?org_id=1")
        assert response.status_code == 200
        assert response.json()["tier"] == "advanced"

        app.dependency_overrides[get_current_user] = lambda: AnonymousUser()
        response = await client.get("/api/v1/analytics/plan-info?org_id=1")
        assert response.status_code == 401
        app.dependency_overrides[get_current_user] = lambda: SimpleNamespace(id=1)

    async def test_ingest_frontend_event_branches(self, client, app):
        app.dependency_overrides[get_current_user] = lambda: AnonymousUser()
        response = await client.post(
            "/api/v1/analytics/events",
            json={"event_name": "page_view", "org_id": 1},
        )
        assert response.status_code == 401
        app.dependency_overrides[get_current_user] = lambda: SimpleNamespace(id=1)

        with _analytics_guard_patches(), patch(
            "src.routers.analytics.track",
            new_callable=AsyncMock,
        ):
            response = await client.post(
                "/api/v1/analytics/events",
                json={"event_name": "bad_event", "org_id": 1},
            )
        assert response.status_code == 400

        with _analytics_guard_patches(), patch(
            "src.routers.analytics.track",
            new_callable=AsyncMock,
        ):
            response = await client.post(
                "/api/v1/analytics/events",
                json={
                    "event_name": "page_view",
                    "org_id": 1,
                    "session_id": "session-1",
                    "properties": {"seconds_spent": 20000, "device_type": "desktop"},
                },
                headers={"cf-ipcountry": "us"},
            )
        assert response.status_code == 200

        with _analytics_guard_patches(), patch(
            "src.routers.analytics.track",
            new_callable=AsyncMock,
        ):
            response = await client.post(
                "/api/v1/analytics/events",
                json={
                    "event_name": "page_view",
                    "org_id": 1,
                    "properties": {"seconds_spent": 0},
                },
            )
        assert response.status_code == 200

        with _analytics_guard_patches(), patch(
            "src.routers.analytics.track",
            new_callable=AsyncMock,
        ):
            response = await client.post(
                "/api/v1/analytics/events",
                json={
                    "event_name": "page_view",
                    "org_id": 1,
                    "properties": {"seconds_spent": "bad"},
                },
            )
        assert response.status_code == 200

    async def test_dashboard_routes(self, client, db_session):
        with _analytics_guard_patches(), patch(
            "src.routers.analytics._enrich_with_metadata",
            new_callable=AsyncMock,
            side_effect=lambda rows, db: rows,
        ), patch(
            "src.routers.analytics._execute_tinybird_query",
            new_callable=AsyncMock,
            return_value={"data": [{"course_uuid": "course_1", "user_id": 1}], "rows": 1, "meta": []},
        ):
            response = await client.get("/api/v1/analytics/dashboard/daily_active_users?org_id=1")
        assert response.status_code == 200
        assert response.json()["rows"] == 1

        user_ns = SimpleNamespace(
            id=1,
            user_uuid="user_1",
            first_name="First",
            last_name="Last",
            username="user1",
            email="user1@test.com",
            avatar_image="",
        )
        scalars_mock = MagicMock()
        scalars_mock.all.return_value = [user_ns]
        result_mock = MagicMock()
        result_mock.scalars.return_value = scalars_mock
        db_session.execute.return_value = result_mock
        with _analytics_guard_patches(), patch(
            "src.security.features_utils.plan_check._check_mode_bypass",
            return_value=None,
        ), patch(
            "src.routers.analytics.get_org_plan",
            new_callable=AsyncMock,
            return_value="starter",
        ), patch(
            "src.routers.analytics.plan_meets_requirement",
            return_value=False,
        ):
            response = await client.get("/api/v1/analytics/dashboard/course_dropoff?org_id=1")
        assert response.status_code == 403

        with _analytics_guard_patches(), patch(
            "src.routers.analytics._enrich_with_metadata",
            new_callable=AsyncMock,
            side_effect=lambda rows, db: rows,
        ), patch(
            "src.routers.analytics._execute_tinybird_query",
            new_callable=AsyncMock,
            return_value={"data": [{"course_uuid": "course_1", "user_id": 1}], "rows": 1, "meta": []},
        ):
            response = await client.get("/api/v1/analytics/dashboard/detail/detail_signups?org_id=1")
        assert response.status_code == 200
        assert response.json()["users"]["1"]["user_uuid"] == "user_1"

        with _analytics_guard_patches():
            response = await client.get("/api/v1/analytics/dashboard/not_a_query?org_id=1")
        assert response.status_code == 404

        with _analytics_guard_patches(), patch(
            "src.routers.analytics.get_org_plan",
            new_callable=AsyncMock,
            return_value="starter",
        ), patch(
            "src.routers.analytics.plan_meets_requirement",
            return_value=False,
        ):
            response = await client.get("/api/v1/analytics/dashboard/db/unknown_query?org_id=1")
        assert response.status_code == 404

        with _analytics_guard_patches(), patch(
            "src.routers.analytics.get_org_plan",
            new_callable=AsyncMock,
            return_value="starter",
        ), patch(
            "src.routers.analytics.plan_meets_requirement",
            return_value=False,
        ):
            response = await client.get("/api/v1/analytics/dashboard/db/grade_distribution?org_id=1")
        assert response.status_code == 403

    async def test_dashboard_db_and_course_routes(self, client, db_session):
        def _make_execute_result(all_result=None, first_result=None):
            scalars = MagicMock()
            scalars.all.return_value = all_result if all_result is not None else []
            scalars.first.return_value = first_result
            result = MagicMock()
            result.scalars.return_value = scalars
            result.all.return_value = all_result if all_result is not None else []
            return result

        db_session.execute.return_value = _make_execute_result(all_result=[(90, 2), (100, 1)])
        with _analytics_guard_patches(), patch(
            "src.routers.analytics.get_org_plan",
            new_callable=AsyncMock,
            return_value="pro",
        ), patch(
            "src.routers.analytics.plan_meets_requirement",
            return_value=True,
        ):
            response = await client.get("/api/v1/analytics/dashboard/db/grade_distribution?org_id=1")
        assert response.status_code == 200
        assert response.json()["data"][0]["grade"] == 90

        with _analytics_guard_patches(), patch(
            "src.routers.analytics.get_org_plan",
            new_callable=AsyncMock,
            return_value="pro",
        ), patch(
            "src.routers.analytics.plan_meets_requirement",
            return_value=True,
        ), patch(
            "src.routers.analytics._enrich_with_metadata",
            new_callable=AsyncMock,
            side_effect=lambda rows, db: rows,
        ), patch(
            "src.routers.analytics._execute_tinybird_query",
            new_callable=AsyncMock,
            return_value={"data": [{"course_uuid": "course_1", "user_id": 1}], "rows": 1, "meta": []},
        ):
            response = await client.get(
                "/api/v1/analytics/dashboard/course/course_overview_stats?org_id=1&course_uuid=course_1"
            )
        assert response.status_code == 200
        assert response.json()["data"][0]["course_uuid"] == "course_1"

        db_session.execute.side_effect = [
            _make_execute_result(first_result=SimpleNamespace(id=100, course_uuid="course_1", name="Course 1")),
            _make_execute_result(all_result=[SimpleNamespace(user_id=7, creation_date="2024-01-01")]),
            _make_execute_result(
                all_result=[
                    SimpleNamespace(
                        id=7,
                        user_uuid="user_7",
                        first_name="User",
                        last_name="Seven",
                        username="user7",
                        email="user7@test.com",
                        avatar_image="",
                    )
                ]
            ),
        ]
        with _analytics_guard_patches(), patch(
            "src.routers.analytics.get_org_plan",
            new_callable=AsyncMock,
            return_value="pro",
        ), patch(
            "src.routers.analytics.plan_meets_requirement",
            return_value=True,
        ), patch(
            "src.routers.analytics._enrich_with_metadata",
            new_callable=AsyncMock,
            side_effect=lambda rows, db: rows,
        ), patch(
            "src.routers.analytics._execute_tinybird_query",
            new_callable=AsyncMock,
            return_value={"data": [], "rows": 0, "meta": []},
        ):
            response = await client.get(
                "/api/v1/analytics/dashboard/course/detail/course_recent_enrollments?org_id=1&course_uuid=course_1"
            )
        assert response.status_code == 200
        assert response.json()["data"][0]["user_id"] == 7

        db_session.execute.side_effect = None
        with _analytics_guard_patches(), patch(
            "src.routers.analytics.get_org_plan",
            new_callable=AsyncMock,
            return_value="starter",
        ), patch(
            "src.routers.analytics.plan_meets_requirement",
            return_value=False,
        ):
            response = await client.get(
                "/api/v1/analytics/dashboard/course/course_overview_stats?org_id=1&course_uuid=course_1"
            )
        assert response.status_code == 403

        with _analytics_guard_patches(), patch(
            "src.routers.analytics.get_org_plan",
            new_callable=AsyncMock,
            return_value="pro",
        ), patch(
            "src.routers.analytics.plan_meets_requirement",
            return_value=True,
        ):
            response = await client.get(
                "/api/v1/analytics/dashboard/course/detail/unknown_detail?org_id=1&course_uuid=course_1"
            )
        assert response.status_code == 404

        with _analytics_guard_patches(), patch(
            "src.routers.analytics.get_org_plan",
            new_callable=AsyncMock,
            return_value="pro",
        ), patch(
            "src.routers.analytics.plan_meets_requirement",
            return_value=True,
        ):
            response = await client.get(
                "/api/v1/analytics/dashboard/course/detail/course_recent_enrollments?org_id=1&course_uuid=bad course"
            )
        assert response.status_code == 400

    async def test_export_json_csv_xlsx_and_errors(self, client):
        with _analytics_guard_patches(), patch(
            "src.routers.analytics._enrich_with_metadata",
            new_callable=AsyncMock,
            side_effect=lambda rows, db: rows,
        ), patch(
            "src.routers.analytics._execute_tinybird_query",
            new_callable=AsyncMock,
            return_value={"data": [{"course_uuid": "course_1", "user_id": 1}], "rows": 1, "meta": []},
        ):
            response = await client.get("/api/v1/analytics/export?org_id=1&format=json")
        assert response.status_code == 200
        assert response.json()["sections"]["overview"]["daily_active_users"]["rows"] == 1

        with _analytics_guard_patches(), patch(
            "src.routers.analytics._enrich_with_metadata",
            new_callable=AsyncMock,
            side_effect=lambda rows, db: rows,
        ), patch(
            "src.routers.analytics._execute_tinybird_query",
            new_callable=AsyncMock,
            return_value={"data": [{"course_uuid": "course_1", "user_id": 1}], "rows": 1, "meta": []},
        ):
            response = await client.get(
                "/api/v1/analytics/export?org_id=1&format=csv"
            )
        assert response.status_code == 200
        assert "# daily_active_users" in response.text

        with _analytics_guard_patches(), patch(
            "src.routers.analytics._enrich_with_metadata",
            new_callable=AsyncMock,
            side_effect=lambda rows, db: rows,
        ), patch(
            "src.routers.analytics._execute_tinybird_query",
            new_callable=AsyncMock,
            return_value={"data": [{"course_uuid": "course_1", "user_id": 1}], "rows": 1, "meta": []},
        ):
            response = await client.get("/api/v1/analytics/export?org_id=1&format=xlsx")
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        with _analytics_guard_patches():
            response = await client.get("/api/v1/analytics/export?org_id=1&format=xml")
        assert response.status_code == 400

    async def test_export_days_and_limit_validation(self, client):
        with _analytics_guard_patches(), patch(
            "src.routers.analytics._execute_tinybird_query",
            new_callable=AsyncMock,
            return_value={"data": [], "rows": 0, "meta": []},
        ):
            response = await client.get("/api/v1/analytics/export?org_id=1&days=99999")
        assert response.status_code == 400
        assert "days" in response.json()["detail"].lower()
    
        with _analytics_guard_patches():
            response = await client.get("/api/v1/analytics/export?org_id=1&limit=0")
        assert response.status_code == 400
    
        with _analytics_guard_patches():
            response = await client.get("/api/v1/analytics/export?org_id=1&limit=-5")
        assert response.status_code == 400

    async def test_dashboard_unknown_and_course_plan_and_query_validation(self, client):
        with _analytics_guard_patches():
            response = await client.get("/api/v1/analytics/dashboard/detail/unknown_detail?org_id=1")
        assert response.status_code == 404

        with _analytics_guard_patches(), patch(
            "src.routers.analytics.get_org_plan",
            new_callable=AsyncMock,
            return_value="starter",
        ), patch(
            "src.routers.analytics.plan_meets_requirement",
            return_value=False,
        ):
            response = await client.get(
                "/api/v1/analytics/dashboard/course/detail/course_recent_enrollments?org_id=1&course_uuid=course_1"
            )
        assert response.status_code == 403

        with _analytics_guard_patches(), patch(
            "src.routers.analytics.get_org_plan",
            new_callable=AsyncMock,
            return_value="pro",
        ), patch(
            "src.routers.analytics.plan_meets_requirement",
            return_value=True,
        ):
            response = await client.get(
                "/api/v1/analytics/dashboard/course/unknown_course_query?org_id=1&course_uuid=course_1"
            )
        assert response.status_code == 404

    async def test_anonymous_routes_return_401(self, client, app):
        app.dependency_overrides[get_current_user] = lambda: AnonymousUser()
        routes = [
            "/api/v1/analytics/dashboard/detail/detail_signups?org_id=1",
            "/api/v1/analytics/dashboard/daily_active_users?org_id=1",
            "/api/v1/analytics/dashboard/db/grade_distribution?org_id=1",
            "/api/v1/analytics/dashboard/course/detail/course_recent_enrollments?org_id=1&course_uuid=course_1",
            "/api/v1/analytics/dashboard/course/course_overview_stats?org_id=1&course_uuid=course_1",
            "/api/v1/analytics/export?org_id=1&format=json&queries=daily_active_users",
            "/api/v1/analytics/plan-info?org_id=1",
        ]

        for route in routes:
            response = await client.get(route)
            assert response.status_code == 401
