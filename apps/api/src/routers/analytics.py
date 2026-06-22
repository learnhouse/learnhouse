import asyncio
import csv
import io
import logging
import re
import json as _json

from collections import defaultdict
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlmodel import select
from sqlmodel.ext.asyncio.session import AsyncSession

from config.config import get_learnhouse_config
from src.core.events.database import get_db_session
from src.db.courses.activities import Activity
from src.db.courses.courses import Course
from src.db.roles import Role
from src.db.trail_runs import TrailRun
from src.db.user_organizations import UserOrganization
from src.db.usergroups import UserGroup
from src.db.usergroup_user import UserGroupUser
from src.db.users import AnonymousUser, APITokenUser, PublicUser, User
from src.security.auth import get_current_user, resolve_acting_user_id
from src.security.features_utils.plan_check import get_org_plan
from src.security.features_utils.plans import plan_meets_requirement
from src.security.superadmin import is_user_superadmin
from src.services.analytics.analytics import track
from src.services.analytics.cache import get_cached_result, set_cached_result
from src.services.analytics.events import ALLOWED_FRONTEND_EVENTS
from src.services.analytics.queries import (
    ADVANCED_QUERIES,
    ALL_QUERIES,
    CORE_QUERIES,
    COURSE_DETAIL_QUERIES,
    COURSE_QUERIES,
    DETAIL_QUERIES,
    DETAIL_ENROLLMENTS,
    DETAIL_COMPLETIONS,
    COURSE_TOP_LEARNERS,
)

import httpx

logger = logging.getLogger(__name__)

router = APIRouter()


# ---------------------------------------------------------------------------
# Request / response models
# ---------------------------------------------------------------------------

class FrontendEvent(BaseModel):
    event_name: str
    org_id: int
    session_id: str = ""
    properties: dict = {}


class PlanInfoResponse(BaseModel):
    tier: str  # "core" | "advanced"
    plan: str


class AnalyticsStatusResponse(BaseModel):
    configured: bool


# ---------------------------------------------------------------------------
# Tinybird read client — lazy singleton
# ---------------------------------------------------------------------------

_read_client: httpx.AsyncClient | None = None


def _get_read_client() -> httpx.AsyncClient | None:
    global _read_client
    if _read_client is not None:
        return _read_client

    config = get_learnhouse_config()
    tb = config.tinybird_config
    if tb is None:
        return None

    _read_client = httpx.AsyncClient(
        base_url=tb.api_url,
        headers={"Authorization": f"Bearer {tb.read_token}"},
        timeout=30.0,
    )
    return _read_client


# ---------------------------------------------------------------------------
# SQL helpers
# ---------------------------------------------------------------------------

_SAFE_COURSE_UUID = re.compile(r"^[a-zA-Z0-9_-]+$")
_MAX_DAYS = 3650   # 10 years hard ceiling
_MAX_SAFE_DAYS = 365  # default ceiling for export


def _validate_course_uuid(course_uuid: str) -> str:
    """Validate course_uuid is safe for interpolation into SQL."""
    if not course_uuid or not _SAFE_COURSE_UUID.match(course_uuid):
        raise HTTPException(status_code=400, detail="Invalid course_uuid")
    if len(course_uuid) > 100:
        raise HTTPException(status_code=400, detail="Invalid course_uuid")
    return course_uuid


def _validate_days(days_param: str | None, default: int) -> int:
    """Parse and clamp the days parameter."""
    try:
        d = int(days_param) if days_param else default
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid 'days' parameter")
    if d < 0 or d > _MAX_DAYS:
        raise HTTPException(status_code=400, detail=f"'days' must be between 0 and {_MAX_DAYS}")
    return d


def _build_sql(
    sql_template: str,
    org_id: int,
    days: int,
    course_uuid: str | None = None,
    limit: int | None = None,
) -> str:
    """
    Interpolate validated parameters into a SQL template.
    org_id and days must already be validated ints.
    course_uuid must already be validated by _validate_course_uuid.
    """
    if not isinstance(org_id, int) or not isinstance(days, int):
        raise HTTPException(status_code=400, detail="Invalid parameter types")

    params: dict = {"org_id": org_id, "days": days}

    if course_uuid is not None:
        if not _SAFE_COURSE_UUID.match(course_uuid) or len(course_uuid) > 100:
            raise HTTPException(status_code=400, detail="Invalid course_uuid")
        params["course_uuid"] = course_uuid

    if limit is not None:
        if not isinstance(limit, int) or limit < 1:
            raise HTTPException(status_code=400, detail="Invalid limit")
        params["limit"] = limit

    return sql_template.format(**params)


def _parse_safe_params(
    org_id: int,
    request: Request,
    default_days: int,
) -> tuple[int, int]:
    """Validate and return (safe_org_id, safe_days)."""
    try:
        safe_org_id = int(org_id)
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid org_id")
    days_param = request.query_params.get("days")
    return safe_org_id, _validate_days(days_param, default_days)

# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

async def _verify_org_membership(
    user_id: int, org_id: int, db_session: AsyncSession
) -> None:
    """Verify the user belongs to the org. Superadmins bypass this check."""
    if await is_user_superadmin(user_id, db_session):
        return
    membership = (await db_session.execute(
        select(UserOrganization).where(
            UserOrganization.user_id == user_id,
            UserOrganization.org_id == org_id,
        )
    )).scalars().first()
    if not membership:
        raise HTTPException(status_code=403, detail="Not a member of this organization")


async def _verify_org_admin(
    user_id: int, org_id: int, db_session: AsyncSession
) -> None:
    """
    Verify the user has admin rights in the org.
    Checks for organizations.action_update permission.
    Superadmins bypass this check.
    """
    if await is_user_superadmin(user_id, db_session):
        return

    membership = (await db_session.execute(
        select(UserOrganization).where(
            UserOrganization.user_id == user_id,
            UserOrganization.org_id == org_id,
        )
    )).scalars().first()
    if not membership:
        raise HTTPException(status_code=403, detail="Admin access required for this organization")

    role = (await db_session.execute(
        select(Role).where(Role.id == membership.role_id)
    )).scalars().first()
    if not role:
        raise HTTPException(status_code=403, detail="Admin access required for this organization")

    if role.rights:
        rights = role.rights
        org_rights = (
            rights.get("organizations")
            if isinstance(rights, dict)
            else getattr(rights, "organizations", None)
        )
        if org_rights:
            has_update = (
                org_rights.get("action_update", False)
                if isinstance(org_rights, dict)
                else getattr(org_rights, "action_update", False)
            )
            if has_update:
                return

    raise HTTPException(status_code=403, detail="Admin access required for this organization")


# ---------------------------------------------------------------------------
# Tinybird query execution with Redis caching
# ---------------------------------------------------------------------------

async def _execute_tinybird_query(
    query_name: str,
    sql: str,
    org_id: int,
    days: int,
    course_id: str | None = None,
    empty_response: dict | None = None,
    bypass_cache: bool = False,
) -> dict:
    """
    Execute a SQL query via Tinybird Query API with Redis caching.
    Returns empty_response (not an error) if the table is not found.
    """
    if empty_response is None:
        empty_response = {"data": [], "rows": 0, "meta": []}

    if not bypass_cache:
        cached = get_cached_result(query_name, org_id, days, course_id)
        if cached is not None:
            return cached

    client = _get_read_client()
    if client is None:
        raise HTTPException(status_code=503, detail="Analytics not configured")

    try:
        resp = await client.post("/v0/sql", content=sql + " FORMAT JSON")
        resp.raise_for_status()
        result = resp.json()
    except httpx.HTTPStatusError as exc:
        error_msg = exc.response.text[:500]
        logger.warning(
            "Tinybird query '%s' failed (%s): %s",
            query_name, exc.response.status_code, error_msg,
        )
        if any(s in error_msg for s in ("UNKNOWN_TABLE", "doesn't exist", "not found")):
            return empty_response
        raise HTTPException(status_code=502, detail="Analytics query failed")
    except Exception as exc:
        logger.warning("Tinybird query '%s' failed: %s", query_name, str(exc)[:500])
        raise HTTPException(status_code=502, detail="Analytics query failed")

    # Sanitize NaN/Inf float values
    rows = result.get("data", [])
    for row in rows:
        for key, val in row.items():
            if isinstance(val, float) and (
                val != val or val == float("inf") or val == float("-inf")
            ):
                row[key] = None

    response = {
        "data": rows,
        "rows": result.get("rows", len(rows)),
        "meta": result.get("meta", []),
    }
    set_cached_result(query_name, org_id, days, response, course_id)
    return response


# ---------------------------------------------------------------------------
# PostgreSQL metadata enrichment
# ---------------------------------------------------------------------------

async def _enrich_with_metadata(
    rows: list[dict], db_session: AsyncSession
) -> list[dict]:
    """Enrich analytics rows with course/activity names from PostgreSQL."""
    if not rows:
        return rows

    course_uuids = {row["course_uuid"] for row in rows if row.get("course_uuid")}
    course_map: dict[str, Course] = {}
    if course_uuids:
        courses = (await db_session.execute(
            select(Course).where(Course.course_uuid.in_(list(course_uuids)))  # type: ignore
        )).scalars().all()
        course_map = {c.course_uuid: c for c in courses}

    activity_uuids = {
        row[key]
        for row in rows
        for key in ("activity_uuid", "last_activity_uuid")
        if row.get(key)
    }
    activity_map: dict[str, Activity] = {}
    if activity_uuids:
        activities = (await db_session.execute(
            select(Activity).where(Activity.activity_uuid.in_(list(activity_uuids)))  # type: ignore
        )).scalars().all()
        activity_map = {a.activity_uuid: a for a in activities}

        activity_course_ids = {a.course_id for a in activities if a.course_id}
        missing_course_ids = activity_course_ids - {c.id for c in course_map.values()}
        if missing_course_ids:
            extra = (await db_session.execute(
                select(Course).where(Course.id.in_(list(missing_course_ids)))  # type: ignore
            )).scalars().all()
            for c in extra:
                course_map[c.course_uuid] = c

    course_id_to_obj = {c.id: c for c in course_map.values()}

    for row in rows:
        if row.get("course_uuid"):
            course = course_map.get(row["course_uuid"])
            if course:
                row["course_name"] = course.name
                row["thumbnail_image"] = course.thumbnail_image or ""

        for key in ("activity_uuid", "last_activity_uuid"):
            if row.get(key):
                activity = activity_map.get(row[key])
                if activity:
                    name_key = "activity_name" if key == "activity_uuid" else "last_activity_name"
                    row[name_key] = activity.name
                    if not row.get("course_uuid") and activity.course_id:
                        parent = course_id_to_obj.get(activity.course_id)
                        if parent:
                            row["course_uuid"] = parent.course_uuid
                            row["course_name"] = parent.name

    return rows


# ---------------------------------------------------------------------------
# XLSX styling constants
# ---------------------------------------------------------------------------

_HDR_FILL = PatternFill("solid", start_color="1F3864")
_HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_ALT_FILL = PatternFill("solid", start_color="EEF2F7")
_NRM_FILL = PatternFill("solid", start_color="FFFFFF")
_GRN_FILL = PatternFill("solid", start_color="C6EFCE")
_RED_FILL = PatternFill("solid", start_color="FFCCCC")
_YLW_FILL = PatternFill("solid", start_color="FFF2CC")  # section header
_SEC_FILL = PatternFill("solid", start_color="D9E1F2")  # sub-section
_CELL_FNT = Font(name="Arial", size=10)
_SEC_FONT = Font(bold=True, name="Arial", size=10, color="1F3864")
_CENTER   = Alignment(horizontal="center", vertical="center")
_LEFT     = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN     = Side(style="thin", color="C0C0C0")
_BRD      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Column names that get green/red highlight based on "Sì"/"No" value
_BOOL_HIGHLIGHT_COLS = {"Corso completato", "Iscritto al corso"}

def _sanitize(val):
    if isinstance(val, (list, dict)):
        return str(val)
    if isinstance(val, float) and (val != val or val in (float("inf"), float("-inf"))):
        return None
    return val

_UUID_FIELDS_TO_STRIP = {"course_uuid", "activity_uuid", "last_activity_uuid"}


def _strip_uuids(rows: list[dict]) -> list[dict]:
    return [
        {k: v for k, v in row.items() if k not in _UUID_FIELDS_TO_STRIP}
        for row in rows
    ]

def _style_header_row(ws, n_cols: int, row: int = 1):
    for ci in range(1, n_cols + 1):
        c = ws.cell(row, ci)
        c.fill, c.font, c.alignment, c.border = _HDR_FILL, _HDR_FONT, _CENTER, _BRD
    ws.row_dimensions[row].height = 22


def _style_section_row(ws, label: str, n_cols: int, row: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row, 1, value=label)
    c.fill = _SEC_FILL
    c.font = _SEC_FONT
    c.alignment = _LEFT
    c.border = _BRD
    ws.row_dimensions[row].height = 18

def _autofit(ws, rows: list[dict], headers: list[str]):
    for ci, h in enumerate(headers, 1):
        col = get_column_letter(ci)
        max_w = max(
            len(str(h)),
            max((len(str(_sanitize(r.get(h)) or "")) for r in rows), default=0),
        )
        ws.column_dimensions[col].width = min(max_w + 4, 40)

def _write_xlsx_sheet(ws, rows: list[dict]) -> None:
    """Write a list of dicts to a worksheet with styled headers. (unchanged)"""
    if not rows:
        ws.append(["No data"])
        return
 
    headers = list(rows[0].keys())
    ws.append(headers)
    _style_header_row(ws, len(headers))
 
    for ri, rec in enumerate(rows, 2):
        base_fill = _ALT_FILL if ri % 2 == 0 else _NRM_FILL
        for ci, k in enumerate(headers, 1):
            val = _sanitize(rec.get(k, ""))
            c = ws.cell(ri, ci, value=val)
            if k in _BOOL_HIGHLIGHT_COLS:
                c.fill = _GRN_FILL if val in ("Yes", "Sì") else _RED_FILL
            else:
                c.fill = base_fill
            c.font = _CELL_FNT
            c.border = _BRD
            c.alignment = _CENTER if ci == 1 else _LEFT
        ws.row_dimensions[ri].height = 18
 
    _autofit(ws, rows, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def _kpi_block(ws, start_row: int, start_col: int, label: str, value, unit: str = ""):
    lc = ws.cell(start_row, start_col, value=label)
    lc.font = Font(bold=True, name="Arial", size=9, color="555555")
    lc.alignment = _CENTER
    vc = ws.cell(start_row + 1, start_col, value=str(value) + (f" {unit}" if unit else ""))
    vc.font = Font(bold=True, name="Arial", size=14, color="1F3864")
    vc.alignment = _CENTER

def _write_overview_sheet(ws, results: dict, generated_at: str):
    ws.column_dimensions["A"].width = 28
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 22

    ws.merge_cells("A1:E1")
    t = ws.cell(1, 1, value="📊 Analytics Overview")
    t.font = Font(bold=True, name="Arial", size=14, color="FFFFFF")
    t.fill = _HDR_FILL
    t.alignment = _CENTER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    s = ws.cell(2, 1, value=f"Generated: {generated_at}")
    s.font = Font(name="Arial", size=9, color="888888")
    s.alignment = _CENTER
    ws.row_dimensions[2].height = 16

    funnel = results.get("enrollment_funnel", {}).get("data", [{}])
    f = funnel[0] if funnel else {}

    growth = results.get("org_growth_trend", {}).get("data", [])
    total_signups   = sum(r.get("signups", 0) or 0 for r in growth)
    total_enrolls   = sum(r.get("enrollments", 0) or 0 for r in growth)
    total_completes = sum(r.get("completions", 0) or 0 for r in growth)

    ws.merge_cells("A4:E4")
    sec = ws.cell(4, 1, value="Key Performance Indicators")
    sec.fill = _SEC_FILL
    sec.font = _SEC_FONT
    sec.alignment = _LEFT
    sec.border = _BRD
    ws.row_dimensions[4].height = 18

    kpis = [
        ("Signups", total_signups),
        ("Enrollments", total_enrolls),
        ("Completions", total_completes),
        ("Completion Rate", (
            f"{round(f.get('completions',0)/f.get('enrollments',1)*100,1)}%"
            if f.get("enrollments") else "N/A"
        )),
    ]
    for ci, (label, value) in enumerate(kpis, 1):
        _kpi_block(ws, 5, ci, label, value)
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 24

    ws.merge_cells("A8:E8")
    s2 = ws.cell(8, 1, value="Enrollment Funnel")
    s2.fill = _SEC_FILL
    s2.font = _SEC_FONT
    s2.alignment = _LEFT
    ws.row_dimensions[8].height = 18

    funnel_headers = ["Page Views", "Course Views", "Enrollments", "Completions", "Enroll→Complete %"]
    for ci, h in enumerate(funnel_headers, 1):
        c = ws.cell(9, ci, h)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = _CENTER
        c.border = _BRD

    enrollments = f.get("enrollments") or 0
    completions = f.get("completions") or 0
    conv = f"{round(completions/enrollments*100,1)}%" if enrollments else "N/A"
    row_vals = [f.get("page_views", ""), f.get("course_views", ""), enrollments, completions, conv]
    for ci, v in enumerate(row_vals, 1):
        c = ws.cell(10, ci, value=v)
        c.fill = _NRM_FILL
        c.font = _CELL_FNT
        c.alignment = _CENTER
        c.border = _BRD
    ws.row_dimensions[9].height = 20
    ws.row_dimensions[10].height = 18

    ws.merge_cells("A12:E12")
    s3 = ws.cell(12, 1, value="Growth Trend (by week)")
    s3.fill = _SEC_FILL
    s3.font = _SEC_FONT
    s3.alignment = _LEFT
    ws.row_dimensions[12].height = 18

    gt_headers = ["Week", "Signups", "Enrollments", "Completions"]
    for ci, h in enumerate(gt_headers, 1):
        c = ws.cell(13, ci, h)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = _CENTER
        c.border = _BRD
    ws.row_dimensions[13].height = 20

    for ri, row in enumerate(growth[-12:], 14):
        fill = _ALT_FILL if ri % 2 == 0 else _NRM_FILL
        for ci, k in enumerate(["week", "signups", "enrollments", "completions"], 1):
            c = ws.cell(ri, ci, value=_sanitize(row.get(k, "")))
            c.fill = fill
            c.font = _CELL_FNT
            c.alignment = _CENTER
            c.border = _BRD
        ws.row_dimensions[ri].height = 16

    ws.freeze_panes = "A3"
 
def _write_learners_sheet(ws, results: dict):
    """
    Sheet 2 — Learners & Growth
    daily_active_users, new_vs_returning, cohort_retention, peak_usage_hours
    """
    ws.column_dimensions["A"].width = 20
    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 16
 
    # Title
    ws.merge_cells("A1:H1")
    t = ws.cell(1, 1, value="Learners & Growth")
    t.font = Font(bold=True, name="Arial", size=13, color="FFFFFF")
    t.fill = _HDR_FILL
    t.alignment = _CENTER
    ws.row_dimensions[1].height = 28
 
    current_row = 3
 
    # --- Daily Active Users ---
    dau = results.get("daily_active_users", {}).get("data", [])
    _style_section_row(ws, "Daily Active Users", 8, current_row)
    current_row += 1
    if dau:
        headers = ["Date", "DAU"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(current_row, ci, h)
            c.fill = _HDR_FILL
            c.font = _HDR_FONT
            c.alignment = _CENTER
            c.border = _BRD
        current_row += 1
        for ri, row in enumerate(dau):
            fill = _ALT_FILL if ri % 2 == 0 else _NRM_FILL
            ws.cell(current_row, 1, value=str(row.get("date", ""))).fill = fill
            ws.cell(current_row, 2, value=row.get("dau", "")).fill = fill
            for ci in range(1, 3):
                ws.cell(current_row, ci).font = _CELL_FNT
                ws.cell(current_row, ci).alignment = _CENTER
                ws.cell(current_row, ci).border = _BRD
            current_row += 1
    else:
        ws.cell(current_row, 1, value="No data").font = _CELL_FNT
        current_row += 1
 
    current_row += 1  # blank row
 
    # --- New vs Returning ---
    nvr = results.get("new_vs_returning", {}).get("data", [])
    _style_section_row(ws, "New vs Returning Users", 8, current_row)
    current_row += 1
    if nvr:
        headers = ["Date", "New Users", "Returning Users"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(current_row, ci, h)
            c.fill = _HDR_FILL
            c.font = _HDR_FONT
            c.alignment = _CENTER
            c.border = _BRD
        current_row += 1
        for ri, row in enumerate(nvr):
            fill = _ALT_FILL if ri % 2 == 0 else _NRM_FILL
            for ci, k in enumerate(["date", "new_users", "returning_users"], 1):
                c = ws.cell(current_row, ci, value=str(row.get(k, "")))
                c.fill = fill
                c.font = _CELL_FNT
                c.alignment = _CENTER
                c.border = _BRD
            current_row += 1
    else:
        ws.cell(current_row, 1, value="No data").font = _CELL_FNT
        current_row += 1
 
    current_row += 1
 
    # --- Cohort Retention ---
    cohort = results.get("cohort_retention", {}).get("data", [])
    _style_section_row(ws, "Cohort Retention (W1 / W2 / W4 / W8)", 8, current_row)
    current_row += 1
    if cohort:
        headers = ["Cohort Week", "Cohort Size", "W1", "W2", "W4", "W8"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(current_row, ci, h)
            c.fill = _HDR_FILL
            c.font = _HDR_FONT
            c.alignment = _CENTER
            c.border = _BRD
        current_row += 1
        for ri, row in enumerate(cohort):
            fill = _ALT_FILL if ri % 2 == 0 else _NRM_FILL
            for ci, k in enumerate(["cohort_week", "cohort_size", "week_1", "week_2", "week_4", "week_8"], 1):
                c = ws.cell(current_row, ci, value=_sanitize(row.get(k, "")))
                c.fill = fill
                c.font = _CELL_FNT
                c.alignment = _CENTER
                c.border = _BRD
            current_row += 1
    else:
        ws.cell(current_row, 1, value="No data").font = _CELL_FNT
        current_row += 1
 
    current_row += 1
 
    # --- Peak Usage Hours (compact: day x hour heatmap summary) ---
    peak = results.get("peak_usage_hours", {}).get("data", [])
    _style_section_row(ws, "Peak Usage Hours (events by day of week)", 8, current_row)
    current_row += 1
    if peak:
        # Aggregate by day_of_week
        by_day: dict[int, int] = defaultdict(int)
        for row in peak:
            by_day[row.get("day_of_week", 0)] += row.get("event_count", 0) or 0
        day_names = {1: "Mon", 2: "Tue", 3: "Wed", 4: "Thu", 5: "Fri", 6: "Sat", 7: "Sun"}
        headers = ["Day", "Total Events"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(current_row, ci, h)
            c.fill = _HDR_FILL
            c.font = _HDR_FONT
            c.alignment = _CENTER
            c.border = _BRD
        current_row += 1
        for ri, (day, count) in enumerate(sorted(by_day.items())):
            fill = _ALT_FILL if ri % 2 == 0 else _NRM_FILL
            ws.cell(current_row, 1, value=day_names.get(day, str(day))).fill = fill
            ws.cell(current_row, 2, value=count).fill = fill
            for ci in range(1, 3):
                ws.cell(current_row, ci).font = _CELL_FNT
                ws.cell(current_row, ci).alignment = _CENTER
                ws.cell(current_row, ci).border = _BRD
            current_row += 1
    else:
        ws.cell(current_row, 1, value="No data").font = _CELL_FNT
 
    ws.freeze_panes = "A2"
 
 
def _write_courses_sheet(ws, results: dict):
    """
    Sheet 3 — Courses & Content
    top_courses, course_rating_by_completion, time_to_completion,
    content_type_effectiveness, course_dropoff, completion_velocity
    """
    ws.column_dimensions["A"].width = 36
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 18
 
    ws.merge_cells("A1:F1")
    t = ws.cell(1, 1, value="Courses & Content")
    t.font = Font(bold=True, name="Arial", size=13, color="FFFFFF")
    t.fill = _HDR_FILL
    t.alignment = _CENTER
    ws.row_dimensions[1].height = 28
 
    current_row = 3
    
    sections = [
        ("Top Courses", "top_courses",
         ["course_name", "views", "enrollments", "completions"],
         ["Course", "Views", "Enrollments", "Completions"]),
        ("Course Effectiveness (Enrollment ≥ 5)", "course_rating_by_completion",
         ["course_name", "enrollments", "completions", "completion_rate"],
         ["Course", "Enrollments", "Completions", "Completion Rate %"]),
        ("Time to Completion (days)", "time_to_completion",
         ["course_name", "median_days", "completions_count"],
         ["Course", "Median Days", "Completions"]),
        ("Content Type Effectiveness", "content_type_effectiveness",
         ["activity_type", "view_count", "completion_count", "completion_rate"],
         ["Type", "Views", "Completions", "Rate %"]),
        ("Course Dropoff Points", "course_dropoff",
         ["course_name", "last_activity_name", "dropoff_count"],
         ["Course", "Last Activity", "Drop-offs"]),
        ("Completion Velocity (avg hrs between activities)", "completion_velocity",
         ["course_name", "avg_hours_between", "transitions"],
         ["Course", "Avg Hours Between", "Transitions"]),
    ]
    
    for section_label, query_key, data_keys, col_labels in sections:
        rows = results.get(query_key, {}).get("data", [])
        _style_section_row(ws, section_label, 6, current_row)
        current_row += 1
 
        for ci, h in enumerate(col_labels, 1):
            c = ws.cell(current_row, ci, h)
            c.fill = _HDR_FILL
            c.font = _HDR_FONT
            c.alignment = _CENTER
            c.border = _BRD
        current_row += 1
 
        if rows:
            for ri, row in enumerate(rows):
                fill = _ALT_FILL if ri % 2 == 0 else _NRM_FILL
                for ci, k in enumerate(data_keys, 1):
                    c = ws.cell(current_row, ci, value=_sanitize(row.get(k, "")))
                    c.fill = fill
                    c.font = _CELL_FNT
                    c.alignment = _LEFT
                    c.border = _BRD
                current_row += 1
        else:
            ws.cell(current_row, 1, value="No data").font = _CELL_FNT
            current_row += 1
 
        current_row += 1  # blank between sections
 
    ws.freeze_panes = "A2"
 

def _build_xlsx_global(results: dict[str, dict]) -> Workbook:
    """
    4-sheet workbook for org-level export:
      1. Overview       — KPI cards + funnel + growth trend
      2. Learners       — DAU, new vs returning, cohort retention, peak hours
      3. Courses        — top courses, effectiveness, time-to-complete, dropoff
      4. Acquisition    — country, device, referrer
    """
    wb = Workbook()
    wb.remove(wb.active)
 
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
 
    ws1 = wb.create_sheet("Overview")
    _write_overview_sheet(ws1, results, generated_at)
 
    ws2 = wb.create_sheet("Learners & Growth")
    _write_learners_sheet(ws2, results)
 
    ws3 = wb.create_sheet("Courses & Content")
    _write_courses_sheet(ws3, results)
 
    return wb


def _build_xlsx_course_users(flat: list[dict]) -> Workbook:
    """
    Course×users workbook:
      1. Summary   — group-level stats
      2. All Users — full flat list
      3. One sheet per group
    """
    wb = Workbook()
    wb.remove(wb.active)
 
    groups: dict[str, list[dict]] = defaultdict(list)
    for row in flat:
        raw = row.get("Groups", "None")
        gs  = [g.strip() for g in raw.split(",")] if raw and raw != "None" else ["No groups"]
        for g in gs:
            if g:
                groups[g].append(row)
 
    # --- Summary sheet ---
    ws_s = wb.create_sheet("Summary")
    ws_s.column_dimensions["A"].width = 28
    for col in ["B", "C", "D", "E"]:
        ws_s.column_dimensions[col].width = 16
 
    ws_s.merge_cells("A1:E1")
    t = ws_s.cell(1, 1, value="Course Export — Summary")
    t.font = Font(bold=True, name="Arial", size=13, color="FFFFFF")
    t.fill = _HDR_FILL
    t.alignment = _CENTER
    ws_s.row_dimensions[1].height = 28
 
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws_s.merge_cells("A2:E2")
    s = ws_s.cell(2, 1, value=f"Generated: {generated_at}  |  Total users: {len(flat)}")
    s.font = Font(name="Arial", size=9, color="888888")
    s.alignment = _CENTER
 
    # KPI row
    enrolled_total  = sum(1 for r in flat if r.get("Subscribed") == "Yes")
    completed_total = sum(1 for r in flat if r.get("Completed Course") == "Yes")
    conv = f"{round(completed_total/enrolled_total*100,1)}%" if enrolled_total else "N/A"
 
    kpis = [
        ("Total Users", len(flat)),
        ("Enrolled", enrolled_total),
        ("Completed", completed_total),
        ("Completion Rate", conv),
        ("Groups", len(groups)),
    ]
    ws_s.merge_cells("A4:E4")
    sec = ws_s.cell(4, 1, value="Key Metrics")
    sec.fill = _SEC_FILL
    sec.font = _SEC_FONT
    sec.alignment = _LEFT
    sec.border = _BRD
    ws_s.row_dimensions[4].height = 18
 
    for ci, (label, value) in enumerate(kpis, 1):
        _kpi_block(ws_s, 5, ci, label, value)
    ws_s.row_dimensions[5].height = 16
    ws_s.row_dimensions[6].height = 24
 
    # Group table
    ws_s.merge_cells("A8:E8")
    sec2 = ws_s.cell(8, 1, value="By Group")
    sec2.fill = _SEC_FILL
    sec2.font = _SEC_FONT
    sec2.alignment = _LEFT
    sec2.border = _BRD
    ws_s.row_dimensions[8].height = 18
 
    hdr = ["Group", "Total Users", "Enrolled", "Completed", "Completion Rate"]
    for ci, h in enumerate(hdr, 1):
        c = ws_s.cell(9, ci, h)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = _CENTER
        c.border = _BRD
    ws_s.row_dimensions[9].height = 20
 
    for ri, (g, members) in enumerate(sorted(groups.items()), 10):
        enrolled  = sum(1 for m in members if m.get("Subscribed") == "Yes")
        completed = sum(1 for m in members if m.get("Completed Course") == "Yes")
        rate = f"{round(completed/enrolled*100,1)}%" if enrolled else "N/A"
        fill = _ALT_FILL if ri % 2 == 0 else _NRM_FILL
        for ci, v in enumerate([g, len(members), enrolled, completed, rate], 1):
            c = ws_s.cell(ri, ci, value=v)
            c.fill = fill
            c.font = _CELL_FNT
            c.alignment = _CENTER
            c.border = _BRD
        ws_s.row_dimensions[ri].height = 16
 
    ws_s.freeze_panes = "A10"
 
    # --- All Users sheet ---
    ws_a = wb.create_sheet("All")
    _write_xlsx_sheet(ws_a, flat)
 
    # --- One sheet per group ---
    for g in sorted(groups.keys()):
        ws = wb.create_sheet(g[:31])
        _write_xlsx_sheet(ws, groups[g])
 
    return wb


# ---------------------------------------------------------------------------
# Course×users data fetcher
# ---------------------------------------------------------------------------

async def _fetch_course_users_flat(
    safe_org_id: int,
    safe_days: int,
    safe_course_uuid: str,
    db_session: AsyncSession,
    for_export: bool = False,
) -> list[dict]:
    """
    Build one row per org user with course enrollment/completion/progress data.
    Fires Tinybird and PostgreSQL queries concurrently where possible.
    """
    limit = 10_000 if for_export else 200
    
    # ── Tinybird queries — run concurrently ───────────────────────────────────
    sql_enrollments = _build_sql(DETAIL_ENROLLMENTS, safe_org_id, safe_days, limit=limit)
    sql_completions = _build_sql(DETAIL_COMPLETIONS, safe_org_id, safe_days, limit=limit)
    sql_top_learners = _build_sql(COURSE_TOP_LEARNERS, safe_org_id, safe_days, safe_course_uuid, limit=limit)

    tb_enrollments_raw, tb_completions_raw, tb_top_learners_raw = await asyncio.gather(
        _execute_tinybird_query("detail_enrollments",  sql_enrollments,  safe_org_id, safe_days,bypass_cache=for_export),
        _execute_tinybird_query("detail_completions",  sql_completions,  safe_org_id, safe_days,bypass_cache=for_export),
        _execute_tinybird_query("course_top_learners", sql_top_learners, safe_org_id, safe_days,
                                course_id=safe_course_uuid,bypass_cache=for_export),
    )

    # Filter enrollments and completions for this specific course
    enrollments: dict[int, str] = {
        int(r["user_id"]): r.get("enrolled_at") or r.get("timestamp", "")
        for r in tb_enrollments_raw.get("data", [])
        if r.get("course_uuid") == safe_course_uuid and r.get("user_id")
    }
    completions: dict[int, str] = {
        int(r["user_id"]): r.get("timestamp", "")
        for r in tb_completions_raw.get("data", [])
        if r.get("course_uuid") == safe_course_uuid and r.get("user_id")
    }
    progress_map: dict[int, dict] = {
        int(r["user_id"]): {
            "Completed Activities": r.get("completions", ""),
            "total_seconds":      r.get("total_seconds_spent", ""),
        }
        for r in tb_top_learners_raw.get("data", [])
        if r.get("user_id")
    }

    # ── PostgreSQL queries — run concurrently ─────────────────────────────────
    memberships_result, uug_result = await asyncio.gather(
        db_session.execute(
            select(UserOrganization).where(UserOrganization.org_id == safe_org_id)
        ),
        db_session.execute(
            select(UserGroupUser)
        ),
    )
    memberships     = memberships_result.scalars().all()
    uug_rows_all    = uug_result.scalars().all()

    user_ids       = [m.user_id for m in memberships]
    membership_map = {m.user_id: m for m in memberships}

    if not user_ids:
        return []

    # Filter usergroup memberships to only org users
    user_ids_set = set(user_ids)
    uug_rows     = [r for r in uug_rows_all if r.user_id in user_ids_set]
    ug_ids       = list({r.usergroup_id for r in uug_rows})

    role_ids = list({m.role_id for m in memberships if m.role_id})

    users_result, roles_result = await asyncio.gather(
        db_session.execute(select(User).where(User.id.in_(user_ids))),   # type: ignore
        db_session.execute(select(Role).where(Role.id.in_(role_ids))),   # type: ignore
    )
    users_map = {u.id: u for u in users_result.scalars().all()}
    roles_map = {r.id: r for r in roles_result.scalars().all()}

    ug_map: dict[int, str] = {}
    if ug_ids:
        ugs_result = await db_session.execute(
            select(UserGroup).where(UserGroup.id.in_(ug_ids))  # type: ignore
        )
        ug_map = {g.id: g.name for g in ugs_result.scalars().all()}

    user_groups: dict[int, list[str]] = defaultdict(list)
    for r in uug_rows:
        name = ug_map.get(r.usergroup_id)
        if name:
            user_groups[r.user_id].append(name)

    # ── Flatten ───────────────────────────────────────────────────────────────
    flat: list[dict] = []
    for uid in user_ids:
        u    = users_map.get(uid)
        m    = membership_map.get(uid)
        role = roles_map.get(m.role_id) if m and m.role_id else None
        an   = progress_map.get(uid, {})
        gs   = user_groups.get(uid, [])

        flat.append({
            "ID":                    uid,
            "Username":              getattr(u, "username", ""),
            "First Name":                  getattr(u, "first_name", ""),
            "Last Name":               getattr(u, "last_name", ""),
            "Email":                 getattr(u, "email", ""),
            "Role":                 getattr(role, "name", "") if role else "",
            "Signup Method":         getattr(u, "signup_method", ""),
            "Verified Mail":      "Sì" if getattr(u, "email_verified", False) else "No",
            "Last Login":          str(getattr(u, "last_login_at", "") or ""),
            "Signup Date":   str(getattr(m, "joined_at", "") or ""),
            "Groups":                ", ".join(gs) if gs else "None",
            "Subscribed":     "Yes" if uid in enrollments or uid in progress_map else "No",
            "Subscription Date": str(enrollments.get(uid, "")),
            "Completed Course":      "Yes" if uid in completions else "No",
            "Completion Date":    str(completions.get(uid, "")),
            "Completed Activities":   an.get("completed_activities", ""),
            "Total Seconds":        an.get("total_seconds", ""),
        })

    return flat


# ---------------------------------------------------------------------------
# Export serializers
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Export constants and JSON helpers
# ---------------------------------------------------------------------------

_QUERY_LABELS = {
    "org_growth_trend":            "Org Growth Trend",
    "enrollment_funnel":           "Enrollment Funnel",
    "daily_active_users":          "Daily Active Users",
    "new_vs_returning":            "New vs Returning Users",
    "cohort_retention":            "Cohort Retention",
    "peak_usage_hours":            "Peak Usage Hours",
    "top_courses":                 "Top Courses",
    "course_rating_by_completion": "Course Effectiveness",
    "time_to_completion":          "Time to Completion",
    "course_dropoff":              "Course Dropoff",
    "content_type_effectiveness":  "Content Type Effectiveness",
    "completion_velocity":         "Completion Velocity",
}

_CSV_ORG_QUERIES = {
    "org_growth_trend",
    "enrollment_funnel",
    "daily_active_users",
    "new_vs_returning",
    "cohort_retention",
    "peak_usage_hours",
    "top_courses",
    "course_rating_by_completion",
    "time_to_completion",
    "course_dropoff",
    "content_type_effectiveness",
    "completion_velocity",
}


def _json_response(data: dict) -> Response:
    body = _json.dumps(data, indent=2, ensure_ascii=False, default=str)
    return Response(content=body, media_type="application/json")

def _json_org_export(results: dict, org_id: int, days: int) -> dict:
    now = datetime.now(timezone.utc).isoformat()
    filtered = {k: v for k, v in results.items() if k in _CSV_ORG_QUERIES}
    return {
        "meta": {
            "org_id":       org_id,
            "days":         days,
            "generated_at": now,
            "query_count":  len(filtered),
        },
        "sections": {
            "overview": {
                "org_growth_trend":  filtered.get("org_growth_trend", {}),
                "enrollment_funnel": filtered.get("enrollment_funnel", {}),
                "daily_active_users": filtered.get("daily_active_users", {}),
            },
            "learners": {
                "new_vs_returning": filtered.get("new_vs_returning", {}),
                "cohort_retention":  filtered.get("cohort_retention", {}),
                "peak_usage_hours":  filtered.get("peak_usage_hours", {}),
            },
            "courses": {
                "top_courses":                 filtered.get("top_courses", {}),
                "course_rating_by_completion": filtered.get("course_rating_by_completion", {}),
                "time_to_completion":          filtered.get("time_to_completion", {}),
                "course_dropoff":              filtered.get("course_dropoff", {}),
                "content_type_effectiveness":  filtered.get("content_type_effectiveness", {}),
                "completion_velocity":         filtered.get("completion_velocity", {}),
            },
        },
    }


def _json_course_export(flat: list[dict], org_id: int, days: int, course_uuid: str) -> dict:
    now = datetime.now(timezone.utc).isoformat()
    enrolled  = sum(1 for r in flat if r.get("Subscribed") == "Yes")
    completed = sum(1 for r in flat if r.get("Completed Course") == "Yes")
    return {
        "meta": {
            "org_id":          org_id,
            "course_uuid":     course_uuid,
            "days":            days,
            "generated_at":    now,
            "total_users":     len(flat),
            "enrolled":        enrolled,
            "completed":       completed,
            "completion_rate": round(completed / enrolled * 100, 1) if enrolled else None,
        },
        "users": flat,
    }

def _flat_to_csv(flat: list[dict]) -> str:
    if not flat:
        return ""

    output = io.StringIO()
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    output.write("# LearnHouse Course Export\n")
    output.write(f"# Generated: {now}\n")
    output.write(f"# Total users: {len(flat)}\n\n")
 
    sanitized = [
        {k: str(v) if isinstance(v, (list, dict)) else v for k, v in row.items()}
        for row in flat
    ]
    all_keys = dict.fromkeys(k for row in sanitized for k in row)
    writer = csv.DictWriter(output, fieldnames=all_keys, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(sanitized)
    return output.getvalue()


def _results_to_csv(results: dict[str, dict]) -> str:
    output = io.StringIO()
 
    # Metadata header
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    output.write("# LearnHouse Analytics Export\n")
    output.write(f"# Generated: {now}\n")
    output.write(f"# Sections: {len(results)}\n\n")
 
    for qname, result in results.items():
        if qname not in _CSV_ORG_QUERIES:
            continue
 
        rows = result.get("data", [])
        output.write(f"# {qname}\n")
 
        if not rows:
            output.write("# No data\n\n")
            continue
 
        sanitized = [
            {k: str(v) if isinstance(v, (list, dict)) else v for k, v in row.items()}
            for row in rows
        ]
        all_keys = dict.fromkeys(k for row in sanitized for k in row)
        writer = csv.DictWriter(output, fieldnames=all_keys, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(sanitized)
        output.write("\n")
 
    return output.getvalue()


def _xlsx_response(wb: Workbook) -> Response:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment"},
    )


def _csv_response(content: str) -> StreamingResponse:
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment"},
    )


# ---------------------------------------------------------------------------
# GET /status
# ---------------------------------------------------------------------------

@router.get(
    "/status",
    summary="Check analytics configuration",
    description="Returns whether the analytics backend (Tinybird) is configured on the server.",
    responses={
        200: {"description": "Analytics configuration status"},
        401: {"description": "Authentication required"},
    },
)
async def analytics_status(
    current_user: PublicUser | AnonymousUser | APITokenUser = Depends(get_current_user),
):
    if isinstance(current_user, AnonymousUser):
        raise HTTPException(status_code=401, detail="Authentication required")
    config = get_learnhouse_config()
    return AnalyticsStatusResponse(configured=config.tinybird_config is not None)


# ---------------------------------------------------------------------------
# POST /events
# ---------------------------------------------------------------------------

@router.post(
    "/events",
    summary="Ingest a frontend analytics event",
    description="Proxies a whitelisted analytics event from the frontend to the analytics backend.",
    responses={
        200: {"description": "Event ingested successfully"},
        400: {"description": "Invalid event name"},
        401: {"description": "Authentication required"},
        403: {"description": "User is not a member of the target organization"},
    },
)
async def ingest_frontend_event(
    body: FrontendEvent,
    request: Request,
    current_user: PublicUser | AnonymousUser | APITokenUser = Depends(get_current_user),
    db_session: AsyncSession = Depends(get_db_session),
):
    if isinstance(current_user, AnonymousUser):
        raise HTTPException(status_code=401, detail="Authentication required")

    await _verify_org_membership(resolve_acting_user_id(current_user), body.org_id, db_session)

    if body.event_name not in ALLOWED_FRONTEND_EVENTS:
        raise HTTPException(status_code=400, detail="Invalid event name")

    ip = request.client.host if request.client else ""

    properties = dict(body.properties)
    if "seconds_spent" in properties:
        try:
            seconds = float(properties["seconds_spent"])
            if seconds <= 0:
                properties.pop("seconds_spent")
            elif seconds > 14400:
                properties["seconds_spent"] = 14400
        except (ValueError, TypeError):
            properties.pop("seconds_spent")

    country = (
        request.headers.get("cf-ipcountry")
        or request.headers.get("x-country-code")
        or ""
    )
    if country and country not in ("XX", "T1"):
        properties.setdefault("country_code", country.upper())

    await track(
        event_name=body.event_name,
        org_id=body.org_id,
        user_id=resolve_acting_user_id(current_user),
        session_id=body.session_id,
        properties=properties,
        source="frontend",
        ip=ip,
    )
    return {"ok": True}


# ---------------------------------------------------------------------------
# GET /dashboard/detail/{query_name}
# ---------------------------------------------------------------------------

@router.get(
    "/dashboard/detail/{query_name}",
    summary="Run a detail dashboard query",
    description="Executes a predefined analytics detail query against Tinybird and enriches results with PostgreSQL metadata. Requires org admin privileges.",
    responses={
        200: {"description": "Detail query results with enriched metadata"},
        401: {"description": "Authentication required"},
        403: {"description": "User is not a member or admin of the organization"},
        404: {"description": "Unknown detail query name"},
        502: {"description": "Upstream analytics query failed"},
        503: {"description": "Analytics backend is not configured"},
    },
)
async def query_dashboard_detail(
    query_name: str,
    org_id: int,
    request: Request,
    current_user: PublicUser | AnonymousUser | APITokenUser = Depends(get_current_user),
    db_session: AsyncSession = Depends(get_db_session),
):
    if isinstance(current_user, AnonymousUser):
        raise HTTPException(status_code=401, detail="Authentication required")

    await _verify_org_membership(resolve_acting_user_id(current_user), org_id, db_session)
    await _verify_org_admin(resolve_acting_user_id(current_user), org_id, db_session)

    if query_name not in DETAIL_QUERIES:
        raise HTTPException(status_code=404, detail="Unknown detail query")

    sql_template, default_days = DETAIL_QUERIES[query_name]
    safe_org_id, safe_days = _parse_safe_params(org_id, request, default_days)
    sql = _build_sql(sql_template, safe_org_id, safe_days)

    ch_result = await _execute_tinybird_query(
        query_name, sql, safe_org_id, safe_days,
        empty_response={"data": [], "users": {}},
    )
    ch_data = await _enrich_with_metadata(ch_result.get("data", []), db_session)

    user_ids = list({int(row["user_id"]) for row in ch_data if row.get("user_id")})
    users_map: dict[int, dict] = {}
    if user_ids:
        users = (await db_session.execute(
            select(User).where(User.id.in_(user_ids))  # type: ignore
        )).scalars().all()
        users_map = {
            u.id: {
                "user_uuid":    u.user_uuid,
                "first_name":   u.first_name,
                "last_name":    u.last_name,
                "username":     u.username,
                "email":        u.email,
                "avatar_image": u.avatar_image,
            }
            for u in users
        }

    return {"data": ch_data, "users": users_map}


# ---------------------------------------------------------------------------
# GET /dashboard/{query_name}
# ---------------------------------------------------------------------------

@router.get(
    "/dashboard/{query_name}",
    summary="Run a dashboard analytics query",
    description="Executes a predefined analytics query via Tinybird. Advanced queries require an Enterprise plan. Requires org admin privileges.",
    responses={
        200: {"description": "Query results with enriched metadata"},
        401: {"description": "Authentication required"},
        403: {"description": "User is not a member/admin, or plan does not permit advanced queries"},
        404: {"description": "Unknown query name"},
        502: {"description": "Upstream analytics query failed"},
        503: {"description": "Analytics backend is not configured"},
    },
)
async def query_dashboard(
    query_name: str,
    org_id: int,
    request: Request,
    current_user: PublicUser | AnonymousUser | APITokenUser = Depends(get_current_user),
    db_session: AsyncSession = Depends(get_db_session),
):
    if isinstance(current_user, AnonymousUser):
        raise HTTPException(status_code=401, detail="Authentication required")

    await _verify_org_membership(resolve_acting_user_id(current_user), org_id, db_session)
    await _verify_org_admin(resolve_acting_user_id(current_user), org_id, db_session)

    if query_name not in ALL_QUERIES:
        raise HTTPException(status_code=404, detail="Unknown query")

    if query_name in ADVANCED_QUERIES:
        from src.security.features_utils.plan_check import _check_mode_bypass
        bypass = _check_mode_bypass("analytics_advanced")
        if bypass is None:
            current_plan = await get_org_plan(org_id, db_session)
            if not plan_meets_requirement(current_plan, "enterprise"):
                raise HTTPException(
                    status_code=403,
                    detail="Advanced analytics requires an Enterprise plan or higher.",
                )

    sql_template, default_days = ALL_QUERIES[query_name]
    safe_org_id, safe_days = _parse_safe_params(org_id, request, default_days)
    sql = _build_sql(sql_template, safe_org_id, safe_days)

    result = await _execute_tinybird_query(query_name, sql, safe_org_id, safe_days)
    result["data"] = await _enrich_with_metadata(result.get("data", []), db_session)
    return result


# ---------------------------------------------------------------------------
# GET /dashboard/db/{query_name}
# ---------------------------------------------------------------------------

@router.get(
    "/dashboard/db/{query_name}",
    summary="Run a PostgreSQL-backed dashboard query",
    description="Executes a predefined PostgreSQL analytics query. Requires org admin privileges and a Pro plan or higher.",
    responses={
        200: {"description": "Query results from PostgreSQL"},
        401: {"description": "Authentication required"},
        403: {"description": "User is not admin or the organization lacks a Pro plan"},
        404: {"description": "Unknown query name"},
    },
)
async def query_dashboard_db(
    query_name: str,
    org_id: int,
    request: Request,
    current_user: PublicUser | AnonymousUser | APITokenUser = Depends(get_current_user),
    db_session: AsyncSession = Depends(get_db_session),
):
    if isinstance(current_user, AnonymousUser):
        raise HTTPException(status_code=401, detail="Authentication required")

    await _verify_org_membership(resolve_acting_user_id(current_user), org_id, db_session)
    await _verify_org_admin(resolve_acting_user_id(current_user), org_id, db_session)

    _DB_QUERIES = frozenset({"grade_distribution"})

    if query_name not in _DB_QUERIES:
        raise HTTPException(status_code=404, detail="Unknown query")

    current_plan = await get_org_plan(org_id, db_session)
    if not plan_meets_requirement(current_plan, "pro"):
        raise HTTPException(
            status_code=403,
            detail="Advanced analytics requires a Pro plan or higher.",
        )

    if query_name == "grade_distribution":
        return await _query_grade_distribution(org_id, db_session)

    raise HTTPException(status_code=404, detail="Unknown query")  # pragma: no cover


async def _query_grade_distribution(org_id: int, db_session: AsyncSession):
    """Grade distribution histogram from PostgreSQL."""
    from sqlalchemy import text
    result = await db_session.execute(
        text("""
            SELECT aus.grade, count(*) AS count
            FROM assignmentusersubmission aus
            JOIN assignment a ON aus.assignment_id = a.id
            JOIN course c ON a.course_id = c.id
            WHERE c.org_id = :org_id
            GROUP BY aus.grade
            ORDER BY aus.grade
        """),
        {"org_id": org_id},
    )
    return {"data": [{"grade": row[0], "count": row[1]} for row in result.all()]}


# ---------------------------------------------------------------------------
# GET /dashboard/course/detail/{query_name}
# ---------------------------------------------------------------------------

@router.get(
    "/dashboard/course/detail/{query_name}",
    summary="Run a course-level detail analytics query",
    description="Executes a predefined course-level detail query via Tinybird with PostgreSQL fallback. Requires org admin privileges and a Pro plan or higher.",
    responses={
        200: {"description": "Course detail query results with enriched metadata"},
        400: {"description": "Invalid query parameter or course_uuid"},
        401: {"description": "Authentication required"},
        403: {"description": "User is not admin or the organization lacks a Pro plan"},
        404: {"description": "Unknown course detail query name"},
        502: {"description": "Upstream analytics query failed"},
        503: {"description": "Analytics backend is not configured"},
    },
)
async def query_course_dashboard_detail(
    query_name: str,
    org_id: int,
    course_uuid: str,
    request: Request,
    current_user: PublicUser | AnonymousUser | APITokenUser = Depends(get_current_user),
    db_session: AsyncSession = Depends(get_db_session),
):
    if isinstance(current_user, AnonymousUser):
        raise HTTPException(status_code=401, detail="Authentication required")

    await _verify_org_membership(resolve_acting_user_id(current_user), org_id, db_session)
    await _verify_org_admin(resolve_acting_user_id(current_user), org_id, db_session)

    current_plan = await get_org_plan(org_id, db_session)
    if not plan_meets_requirement(current_plan, "pro"):
        raise HTTPException(status_code=403, detail="Course analytics requires a Pro plan or higher.")

    if query_name not in COURSE_DETAIL_QUERIES:
        raise HTTPException(status_code=404, detail="Unknown course detail query")

    safe_course_uuid = _validate_course_uuid(course_uuid)
    sql_template, default_days = COURSE_DETAIL_QUERIES[query_name]
    safe_org_id, safe_days = _parse_safe_params(org_id, request, default_days)
    sql = _build_sql(sql_template, safe_org_id, safe_days, safe_course_uuid)

    ch_result = await _execute_tinybird_query(
        query_name, sql, safe_org_id, safe_days,
        course_id=safe_course_uuid,
        empty_response={"data": [], "users": {}},
    )
    ch_data = ch_result.get("data", [])

    # Fallback to PostgreSQL TrailRun for enrollment data
    if query_name == "course_recent_enrollments" and not ch_data:
        course_obj = (await db_session.execute(
            select(Course).where(Course.course_uuid == safe_course_uuid)
        )).scalars().first()
        if course_obj:
            trail_runs = (await db_session.execute(
                select(TrailRun)
                .where(TrailRun.course_id == course_obj.id, TrailRun.org_id == safe_org_id)
                .order_by(TrailRun.id.desc())  # type: ignore
                .limit(50)
            )).scalars().all()
            ch_data = [
                {
                    "user_id": tr.user_id,
                    "enrolled_at": tr.creation_date,
                }
                for tr in trail_runs
            ]

    # Enrich with course/activity metadata from PostgreSQL
    ch_data = await _enrich_with_metadata(ch_data, db_session)

    user_ids = list({int(row["user_id"]) for row in ch_data if row.get("user_id")})
    users_map: dict[int, dict] = {}
    if user_ids:
        users = (await db_session.execute(
            select(User).where(User.id.in_(user_ids))  # type: ignore
        )).scalars().all()
        users_map = {
            u.id: {
                "user_uuid":    u.user_uuid,
                "first_name":   u.first_name,
                "last_name":    u.last_name,
                "username":     u.username,
                "email":        u.email,
                "avatar_image": u.avatar_image,
            }
            for u in users
        }

    return {"data": ch_data, "users": users_map}


# ---------------------------------------------------------------------------
# GET /dashboard/course/{query_name}
# ---------------------------------------------------------------------------

@router.get(
    "/dashboard/course/{query_name}",
    summary="Run a course-level analytics query",
    description="Executes a predefined course-level analytics query via Tinybird. Requires org admin privileges and a Pro plan or higher.",
    responses={
        200: {"description": "Course query results with enriched metadata"},
        400: {"description": "Invalid query parameter or course_uuid"},
        401: {"description": "Authentication required"},
        403: {"description": "User is not admin or the organization lacks a Pro plan"},
        404: {"description": "Unknown course query name"},
        502: {"description": "Upstream analytics query failed"},
        503: {"description": "Analytics backend is not configured"},
    },
)
async def query_course_dashboard(
    query_name: str,
    org_id: int,
    course_uuid: str,
    request: Request,
    current_user: PublicUser | AnonymousUser | APITokenUser = Depends(get_current_user),
    db_session: AsyncSession = Depends(get_db_session),
):
    if isinstance(current_user, AnonymousUser):
        raise HTTPException(status_code=401, detail="Authentication required")

    await _verify_org_membership(resolve_acting_user_id(current_user), org_id, db_session)
    await _verify_org_admin(resolve_acting_user_id(current_user), org_id, db_session)

    current_plan = await get_org_plan(org_id, db_session)
    if not plan_meets_requirement(current_plan, "pro"):
        raise HTTPException(status_code=403, detail="Course analytics requires a Pro plan or higher.")

    if query_name not in COURSE_QUERIES:
        raise HTTPException(status_code=404, detail="Unknown course query")

    safe_course_uuid = _validate_course_uuid(course_uuid)
    sql_template, default_days = COURSE_QUERIES[query_name]
    safe_org_id, safe_days = _parse_safe_params(org_id, request, default_days)
    sql = _build_sql(sql_template, safe_org_id, safe_days, safe_course_uuid)

    result = await _execute_tinybird_query(
        query_name, sql, safe_org_id, safe_days, course_id=safe_course_uuid,
    )
    result["data"] = await _enrich_with_metadata(result.get("data", []), db_session)
    return result


# ---------------------------------------------------------------------------
# GET /export
# ---------------------------------------------------------------------------

@router.get(
    "/export",
    summary="Export analytics data",
    description=(
        "Exports analytics data in JSON, CSV, or XLSX format. "
        "When course_uuid is provided, returns one row per user with enrollment/completion data. "
        "Otherwise returns aggregated org-level analytics (CORE + ADVANCED queries). "
        "The 'queries' parameter is ignored — query sets are hardcoded per context. "
        "Requires org admin privileges."
    ),
    responses={
        200: {"description": "Analytics data exported as JSON, CSV, or XLSX"},
        400: {"description": "Invalid format or parameters"},
        401: {"description": "Authentication required"},
        403: {"description": "User is not admin of the organization"},
        502: {"description": "Upstream analytics query failed"},
        503: {"description": "Analytics backend is not configured"},
    },
)
async def export_analytics(
    org_id: int,
    request: Request,
    current_user: PublicUser | AnonymousUser | APITokenUser = Depends(get_current_user),
    db_session: AsyncSession = Depends(get_db_session),
):
    if isinstance(current_user, AnonymousUser):
        raise HTTPException(status_code=401, detail="Authentication required")

    # ── Valida i parametri prima dei check di autorizzazione ─────────────────
    fmt = request.query_params.get("format", "json")
    if fmt not in ("json", "csv", "xlsx"):
        raise HTTPException(status_code=400, detail="format must be 'json', 'csv', or 'xlsx'")

    safe_days = _validate_days(request.query_params.get("days"), _MAX_SAFE_DAYS)

    limit_param = request.query_params.get("limit")
    if limit_param is not None:
        try:
            limit_val = int(limit_param)
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="Invalid 'limit' parameter")
        if limit_val < 1:
            raise HTTPException(status_code=400, detail="'limit' must be >= 1")

    await _verify_org_membership(resolve_acting_user_id(current_user), org_id, db_session)
    await _verify_org_admin(resolve_acting_user_id(current_user), org_id, db_session)

    course_uuid = request.query_params.get("course_uuid")
    safe_course_uuid = _validate_course_uuid(course_uuid) if course_uuid else None

    safe_org_id = int(org_id)

    # ── Course export: one row per user ───────────────────────────────────────
    if safe_course_uuid:
        flat = await _fetch_course_users_flat(safe_org_id, safe_days, safe_course_uuid, db_session, for_export=True)

        if fmt == "json":
            return _json_response(_json_course_export(flat, safe_org_id, safe_days, safe_course_uuid))

        if fmt == "csv":
            return _csv_response(_flat_to_csv(flat))

        # xlsx
        return _xlsx_response(_build_xlsx_course_users(flat))

    # ── Global export: aggregated org analytics ───────────────────────────────
    export_queries = {
        k: v for k, v in {**CORE_QUERIES, **ADVANCED_QUERIES}.items()
        if k in _CSV_ORG_QUERIES
    }
    results: dict[str, dict] = {}

    for qname, (sql_template, default_days) in export_queries.items():
        d   = safe_days if safe_days else default_days
        sql = _build_sql(sql_template, safe_org_id, d)
        try:
            result = await _execute_tinybird_query(qname, sql, safe_org_id, d)
            result["data"] = await _enrich_with_metadata(result.get("data", []), db_session)
            result["data"] = _strip_uuids(result["data"])
        except HTTPException:
            result = {"data": [], "rows": 0, "meta": []}
        results[qname] = result

    if fmt == "json":
        return _json_response(_json_org_export(results, safe_org_id, safe_days))

    if fmt == "csv":
        return _csv_response(_results_to_csv(results))

    # xlsx
    return _xlsx_response(_build_xlsx_global(results))


# ---------------------------------------------------------------------------
# GET /plan-info
# ---------------------------------------------------------------------------

@router.get(
    "/plan-info",
    summary="Get analytics plan tier for an organization",
    description="Returns the analytics tier (core or advanced) based on the org's current plan.",
    responses={
        200: {"description": "Analytics plan tier information"},
        401: {"description": "Authentication required"},
        403: {"description": "User is not a member of the target organization"},
    },
)
async def get_plan_info(
    org_id: int,
    current_user: PublicUser | AnonymousUser | APITokenUser = Depends(get_current_user),
    db_session: AsyncSession = Depends(get_db_session),
):
    if isinstance(current_user, AnonymousUser):
        raise HTTPException(status_code=401, detail="Authentication required")

    await _verify_org_membership(resolve_acting_user_id(current_user), org_id, db_session)

    current_plan = await get_org_plan(org_id, db_session)
    tier = "advanced" if plan_meets_requirement(current_plan, "pro") else "core"
    return PlanInfoResponse(tier=tier, plan=current_plan)